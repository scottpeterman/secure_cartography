"""
Device Poll Dialog - SNMP fingerprinting from the map viewer

Polls a selected device using SNMP to identify vendor, model, OS via Recog patterns
and OUI lookup. Supports direct SNMP or proxy mode for remote access.

Proxy v2: Ticket-based async polling with progress tracking and cancellation.
"""

import json
import ssl
import urllib.request
import urllib.error
from pathlib import Path
from typing import Optional

from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QTextEdit, QGroupBox, QFormLayout,
    QProgressBar, QTabWidget, QWidget,
    QTableWidget, QTableWidgetItem, QHeaderView, QComboBox,
    QCheckBox, QScrollArea, QFrame
)

from sc2.ui.themes import ThemeColors, ThemeManager

# Import from the new poll_worker module
from .poll_worker import PollWorker, PollResult


# =============================================================================
# Proxy Settings Persistence
# =============================================================================

class ProxySettings:
    """Manages proxy settings persistence"""

    SETTINGS_FILE = Path.home() / '.scng' / 'proxy_settings.json'

    def __init__(self):
        self.enabled: bool = False
        self.url: str = "http://localhost:8899"
        self.api_key: str = ""
        self.load()

    def load(self):
        """Load settings from file"""
        if self.SETTINGS_FILE.exists():
            try:
                with open(self.SETTINGS_FILE, 'r') as f:
                    data = json.load(f)
                    self.enabled = data.get('enabled', False)
                    self.url = data.get('url', "http://localhost:8899")
                    self.api_key = data.get('api_key', "")
            except (json.JSONDecodeError, IOError):
                pass

    def save(self):
        """Save settings to file"""
        self.SETTINGS_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(self.SETTINGS_FILE, 'w') as f:
            json.dump({
                'enabled': self.enabled,
                'url': self.url,
                'api_key': self.api_key
            }, f, indent=2)


# =============================================================================
# Device Poll Dialog
# =============================================================================

class DevicePollDialog(QDialog):
    """Dialog for SNMP polling a device from the map viewer"""

    # Emitted when poll completes with data that could update the node
    node_update_available = pyqtSignal(dict)

    def __init__(
            self,
            ip: str,
            hostname: str = "",
            theme_manager: Optional[ThemeManager] = None,
            parent=None
    ):
        super().__init__(parent)

        self.ip = ip
        self.hostname = hostname or ip
        self.theme_manager = theme_manager
        self._worker: Optional[PollWorker] = None
        self._result: Optional[PollResult] = None

        # Data directory for Recog/OUI files
        self._data_dir = Path.home() / '.scng' / 'fingerprint_data'

        # Load proxy settings
        self._proxy_settings = ProxySettings()

        self.setWindowTitle(f"Poll Device - {self.hostname}")
        self.setMinimumSize(550, 450)
        self.resize(700, 600)

        self._setup_ui()

        if theme_manager:
            self._apply_theme(theme_manager.theme)

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)

        # Scroll area for settings
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.Shape.NoFrame)
        scroll_area.setMaximumHeight(280)  # Limit height so tabs get space

        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        scroll_layout.setContentsMargins(0, 0, 0, 0)
        scroll_layout.setSpacing(8)

        # Target info
        target_group = QGroupBox("Target")
        target_layout = QFormLayout(target_group)
        target_layout.addRow("IP Address:", QLabel(self.ip))
        target_layout.addRow("Hostname:", QLabel(self.hostname))
        scroll_layout.addWidget(target_group)

        # SNMP Settings
        snmp_group = QGroupBox("SNMP Settings")
        snmp_layout = QFormLayout(snmp_group)

        self._community_input = QLineEdit()
        self._community_input.setEchoMode(QLineEdit.EchoMode.Password)
        self._community_input.setText("public")
        snmp_layout.addRow("Community:", self._community_input)

        self._version_combo = QComboBox()
        self._version_combo.addItems(['v1', 'v2c'])
        self._version_combo.setCurrentText('v2c')
        snmp_layout.addRow("Version:", self._version_combo)

        # Download data files button
        download_btn = QPushButton("üì• Download Data Files")
        download_btn.clicked.connect(self._on_download_data)
        snmp_layout.addRow("", download_btn)

        scroll_layout.addWidget(snmp_group)

        # Proxy Settings
        proxy_group = QGroupBox("Proxy Settings")
        proxy_layout = QFormLayout(proxy_group)

        self._proxy_checkbox = QCheckBox("Use SNMP Proxy")
        self._proxy_checkbox.setChecked(self._proxy_settings.enabled)
        self._proxy_checkbox.toggled.connect(self._on_proxy_toggled)
        proxy_layout.addRow("", self._proxy_checkbox)

        self._proxy_url_input = QLineEdit()
        self._proxy_url_input.setText(self._proxy_settings.url)
        self._proxy_url_input.setPlaceholderText("http://jumpbox:8899")
        self._proxy_url_input.setEnabled(self._proxy_settings.enabled)
        proxy_layout.addRow("Proxy URL:", self._proxy_url_input)

        self._proxy_key_input = QLineEdit()
        self._proxy_key_input.setText(self._proxy_settings.api_key)
        self._proxy_key_input.setPlaceholderText("UUID4 from proxy console")
        self._proxy_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self._proxy_key_input.setEnabled(self._proxy_settings.enabled)
        proxy_layout.addRow("API Key:", self._proxy_key_input)

        # Test proxy button
        self._test_proxy_btn = QPushButton("üîó Test Proxy")
        self._test_proxy_btn.clicked.connect(self._on_test_proxy)
        self._test_proxy_btn.setEnabled(self._proxy_settings.enabled)
        proxy_layout.addRow("", self._test_proxy_btn)

        scroll_layout.addWidget(proxy_group)

        scroll_area.setWidget(scroll_content)
        layout.addWidget(scroll_area)

        # Progress
        self._progress_bar = QProgressBar()
        self._progress_bar.setRange(0, 0)  # Indeterminate
        self._progress_bar.setVisible(False)
        layout.addWidget(self._progress_bar)

        self._status_label = QLabel("Ready to poll")
        layout.addWidget(self._status_label)

        # Results tabs
        self._results_tabs = QTabWidget()

        # Summary tab
        summary_widget = QWidget()
        summary_layout = QVBoxLayout(summary_widget)
        self._summary_text = QTextEdit()
        self._summary_text.setReadOnly(True)
        summary_layout.addWidget(self._summary_text)
        self._results_tabs.addTab(summary_widget, "Summary")

        # Interfaces tab
        interfaces_widget = QWidget()
        interfaces_layout = QVBoxLayout(interfaces_widget)
        self._interfaces_table = QTableWidget()
        self._interfaces_table.setColumnCount(3)
        self._interfaces_table.setHorizontalHeaderLabels(['INTERFACE', 'MAC ADDRESS', 'VENDOR'])
        self._interfaces_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        interfaces_layout.addWidget(self._interfaces_table)
        self._results_tabs.addTab(interfaces_widget, "Interfaces")

        # ARP tab
        arp_widget = QWidget()
        arp_layout = QVBoxLayout(arp_widget)
        self._arp_table = QTableWidget()
        self._arp_table.setColumnCount(3)
        self._arp_table.setHorizontalHeaderLabels(['IP ADDRESS', 'MAC ADDRESS', 'VENDOR'])
        self._arp_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        arp_layout.addWidget(self._arp_table)
        self._results_tabs.addTab(arp_widget, "ARP Table")

        # Raw data tab
        raw_widget = QWidget()
        raw_layout = QVBoxLayout(raw_widget)
        self._raw_text = QTextEdit()
        self._raw_text.setReadOnly(True)
        self._raw_text.setFontFamily("monospace")
        raw_layout.addWidget(self._raw_text)
        self._results_tabs.addTab(raw_widget, "Raw Data")

        layout.addWidget(self._results_tabs)

        # Buttons
        button_layout = QHBoxLayout()

        self._poll_btn = QPushButton("üîç Poll Device")
        self._poll_btn.clicked.connect(self._on_poll)
        button_layout.addWidget(self._poll_btn)

        self._cancel_btn = QPushButton("‚õî Cancel")
        self._cancel_btn.clicked.connect(self._on_cancel)
        self._cancel_btn.setEnabled(False)
        self._cancel_btn.setVisible(False)
        button_layout.addWidget(self._cancel_btn)

        self._update_btn = QPushButton("üìù Update Node")
        self._update_btn.clicked.connect(self._on_update_node)
        self._update_btn.setEnabled(False)
        button_layout.addWidget(self._update_btn)

        self._export_btn = QPushButton("üìä Export Excel")
        self._export_btn.clicked.connect(self._on_export_excel)
        self._export_btn.setEnabled(False)
        button_layout.addWidget(self._export_btn)

        button_layout.addStretch()

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)

        layout.addLayout(button_layout)

    def _on_proxy_toggled(self, checked: bool):
        """Handle proxy checkbox toggle"""
        self._proxy_url_input.setEnabled(checked)
        self._proxy_key_input.setEnabled(checked)
        self._test_proxy_btn.setEnabled(checked)

        # Save settings
        self._proxy_settings.enabled = checked
        self._proxy_settings.save()

    def _on_test_proxy(self):
        """Test proxy connectivity"""
        proxy_url = self._proxy_url_input.text().strip().rstrip('/')

        if not proxy_url:
            self._status_label.setText("Enter proxy URL first")
            return

        self._status_label.setText("Testing proxy connection...")

        try:
            req = urllib.request.Request(f"{proxy_url}/health")
            with urllib.request.urlopen(req, timeout=5) as response:
                data = json.loads(response.read().decode('utf-8'))

            if data.get('status') == 'ok':
                version = data.get('version', 'unknown')
                active = data.get('active_jobs', 0)
                max_conc = data.get('max_concurrent', '?')
                self._status_label.setText(
                    f"‚úì Proxy OK - v{version}, {active}/{max_conc} active jobs"
                )
            else:
                self._status_label.setText("Proxy responded but status not OK")

        except urllib.error.URLError as e:
            self._status_label.setText(f"Cannot connect: {e.reason}")
        except Exception as e:
            self._status_label.setText(f"Error: {e}")

    def _on_poll(self):
        """Start polling the device"""
        if self._worker and self._worker.isRunning():
            return

        community = self._community_input.text().strip()
        if not community:
            self._status_label.setText("Enter community string")
            return

        version = self._version_combo.currentText().replace('v', '')

        # Save proxy settings before poll
        self._proxy_settings.enabled = self._proxy_checkbox.isChecked()
        self._proxy_settings.url = self._proxy_url_input.text().strip()
        self._proxy_settings.api_key = self._proxy_key_input.text().strip()
        self._proxy_settings.save()

        # Validate proxy settings if enabled
        if self._proxy_settings.enabled:
            if not self._proxy_settings.url:
                self._status_label.setText("Enter proxy URL")
                return
            if not self._proxy_settings.api_key:
                self._status_label.setText("Enter proxy API key")
                return

        self._poll_btn.setEnabled(False)
        self._cancel_btn.setEnabled(True)
        self._cancel_btn.setVisible(True)
        self._progress_bar.setVisible(True)
        self._status_label.setText("Starting poll...")

        # Clear previous results
        self._clear_results()

        self._worker = PollWorker(
            ip=self.ip,
            community=community,
            data_dir=self._data_dir,
            use_proxy=self._proxy_settings.enabled,
            proxy_url=self._proxy_settings.url,
            proxy_api_key=self._proxy_settings.api_key,
            snmp_version=version
        )
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_poll_finished)
        self._worker.start()

    def _on_cancel(self):
        """Cancel the current poll"""
        if self._worker:
            self._status_label.setText("Cancelling...")
            self._cancel_btn.setEnabled(False)
            self._worker.cancel()

    def _on_progress(self, message: str):
        self._status_label.setText(message)

    def _on_poll_finished(self, result: PollResult):
        self._result = result
        self._poll_btn.setEnabled(True)
        self._cancel_btn.setEnabled(False)
        self._cancel_btn.setVisible(False)
        self._progress_bar.setVisible(False)
        self._worker = None

        if not result.success:
            self._status_label.setText(f"Poll failed: {result.error}")
            return

        # Build status with timing info if available
        status_msg = "Poll complete"
        if result.timing:
            total = result.timing.get('total', 0)
            status_msg = f"Poll complete ({total:.1f}s)"

        self._status_label.setText(status_msg)
        self._update_btn.setEnabled(True)
        self._export_btn.setEnabled(True)

        # Populate tabs
        self._populate_summary(result)
        self._populate_interfaces(result)
        self._populate_arp(result)
        self._populate_raw(result)

    def _clear_results(self):
        """Clear all result displays"""
        self._summary_text.clear()
        self._interfaces_table.setRowCount(0)
        self._arp_table.setRowCount(0)
        self._raw_text.clear()

    def _populate_summary(self, result: PollResult):
        lines = []

        # SNMP data
        for key, value in result.snmp_data.items():
            lines.append(f"{key}: {value}")

        lines.append("")

        # Fingerprint matches
        if result.recog_matches:
            lines.append("=== Fingerprint Matches ===")
            for match in result.recog_matches:
                lines.append(f"\n{match.matched}")
                for k, v in match.params.items():
                    lines.append(f"  {k}: {v}")

        # Timing info
        if result.timing:
            lines.append("")
            lines.append("=== Timing ===")
            for phase, duration in result.timing.items():
                lines.append(f"  {phase}: {duration:.2f}s")

        self._summary_text.setPlainText('\n'.join(lines))

    def _populate_interfaces(self, result: PollResult):
        self._interfaces_table.setRowCount(len(result.interfaces))

        for i, iface in enumerate(result.interfaces):
            self._interfaces_table.setItem(i, 0, QTableWidgetItem(iface.get('description', '')))

            mac = iface.get('mac', '')
            self._interfaces_table.setItem(i, 1, QTableWidgetItem(mac))

            vendor = ''
            if mac and mac in result.oui_lookups:
                vendor = result.oui_lookups[mac].get('manufacturer', '')
            self._interfaces_table.setItem(i, 2, QTableWidgetItem(vendor))

    def _populate_arp(self, result: PollResult):
        self._arp_table.setRowCount(len(result.arp_table))

        for i, entry in enumerate(result.arp_table):
            self._arp_table.setItem(i, 0, QTableWidgetItem(entry.get('ip', '')))

            mac = entry.get('mac', '')
            self._arp_table.setItem(i, 1, QTableWidgetItem(mac))

            vendor = ''
            if mac and mac in result.oui_lookups:
                vendor = result.oui_lookups[mac].get('manufacturer', '')
            self._arp_table.setItem(i, 2, QTableWidgetItem(vendor))

    def _populate_raw(self, result: PollResult):
        raw_data = {
            'ip': result.ip,
            'snmp_data': result.snmp_data,
            'recog_matches': [
                {'matched': m.matched, 'params': m.params}
                for m in result.recog_matches
            ],
            'oui_lookups': result.oui_lookups,
            'interfaces': result.interfaces,
            'arp_table': result.arp_table,
            'timing': result.timing,
        }

        self._raw_text.setPlainText(json.dumps(raw_data, indent=2))

    def _on_update_node(self):
        """Emit signal with data to update the node"""
        if not self._result or not self._result.success:
            return

        # Build update data from results
        update_data = {
            'id': self.hostname,
            'ip': self.ip,
        }

        # Extract platform from Recog matches
        for match in self._result.recog_matches:
            params = match.params
            if 'os.vendor' in params:
                update_data['vendor'] = params['os.vendor']
            if 'os.product' in params:
                update_data['platform'] = params['os.product']
            if 'os.version' in params:
                update_data['version'] = params['os.version']
            if 'os.device' in params:
                update_data['device_type'] = params['os.device']

        # Use sysDescr as fallback platform
        if 'platform' not in update_data and 'sysDescr' in self._result.snmp_data:
            update_data['platform'] = self._result.snmp_data['sysDescr'][:100]

        self.node_update_available.emit(update_data)
        self._status_label.setText("Node update sent")

    def _on_export_excel(self):
        """Export results to Excel"""
        if not self._result:
            return

        from PyQt6.QtWidgets import QFileDialog

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export Results",
            f"{self.hostname}_poll.xlsx",
            "Excel Files (*.xlsx)"
        )

        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            wb = Workbook()

            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # === Summary Sheet ===
            ws_summary = wb.active
            ws_summary.title = "Summary"

            row = 1
            ws_summary.cell(row=row, column=1, value="Property").font = header_font
            ws_summary.cell(row=row, column=1).fill = header_fill
            ws_summary.cell(row=row, column=2, value="Value").font = header_font
            ws_summary.cell(row=row, column=2).fill = header_fill

            row += 1
            for key, value in self._result.snmp_data.items():
                ws_summary.cell(row=row, column=1, value=key).border = thin_border
                ws_summary.cell(row=row, column=2, value=value).border = thin_border
                row += 1

            # Fingerprints
            if self._result.recog_matches:
                row += 1
                ws_summary.cell(row=row, column=1, value="Fingerprint Matches").font = Font(bold=True)
                row += 1
                for match in self._result.recog_matches:
                    ws_summary.cell(row=row, column=1, value=match.matched).border = thin_border
                    row += 1
                    for k, v in match.params.items():
                        ws_summary.cell(row=row, column=1, value=f"  {k}").border = thin_border
                        ws_summary.cell(row=row, column=2, value=v).border = thin_border
                        row += 1

            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 60

            # === Interfaces Sheet ===
            ws_interfaces = wb.create_sheet("Interfaces")

            headers = ['Interface', 'MAC Address', 'Vendor']
            for col, header in enumerate(headers, 1):
                cell = ws_interfaces.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, iface in enumerate(self._result.interfaces, 2):
                ws_interfaces.cell(row=row_idx, column=1, value=iface.get('description', '')).border = thin_border

                mac = iface.get('mac', '')
                ws_interfaces.cell(row=row_idx, column=2, value=mac).border = thin_border

                vendor = ''
                if mac and mac in self._result.oui_lookups:
                    vendor = self._result.oui_lookups[mac].get('manufacturer', '')
                ws_interfaces.cell(row=row_idx, column=3, value=vendor).border = thin_border

            ws_interfaces.column_dimensions['A'].width = 30
            ws_interfaces.column_dimensions['B'].width = 20
            ws_interfaces.column_dimensions['C'].width = 40

            # === ARP Table Sheet ===
            ws_arp = wb.create_sheet("ARP Table")

            headers = ['IP Address', 'MAC Address', 'Vendor']
            for col, header in enumerate(headers, 1):
                cell = ws_arp.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_idx, entry in enumerate(self._result.arp_table, 2):
                ws_arp.cell(row=row_idx, column=1, value=entry.get('ip', '')).border = thin_border

                mac = entry.get('mac', '')
                ws_arp.cell(row=row_idx, column=2, value=mac).border = thin_border

                vendor = ''
                if mac and mac in self._result.oui_lookups:
                    vendor = self._result.oui_lookups[mac].get('manufacturer', '')
                ws_arp.cell(row=row_idx, column=3, value=vendor).border = thin_border

            ws_arp.column_dimensions['A'].width = 18
            ws_arp.column_dimensions['B'].width = 20
            ws_arp.column_dimensions['C'].width = 40

            # Save
            wb.save(filepath)
            self._status_label.setText(f"Exported to {Path(filepath).name}")

        except Exception as e:
            self._status_label.setText(f"Export failed: {e}")

    def _on_download_data(self):
        """Download fingerprint data files"""
        self._status_label.setText("Downloading data files...")
        self._progress_bar.setVisible(True)

        try:
            self._data_dir.mkdir(parents=True, exist_ok=True)
            recog_dir = self._data_dir / 'recog'
            recog_dir.mkdir(exist_ok=True)

            # Create SSL context that doesn't verify certificates
            ssl_context = ssl.create_default_context()
            ssl_context.check_hostname = False
            ssl_context.verify_mode = ssl.CERT_NONE

            files = [
                (recog_dir / 'snmp_sysdescr.xml',
                 'https://raw.githubusercontent.com/rapid7/recog/main/xml/snmp_sysdescr.xml'),
                (recog_dir / 'snmp_sysobjid.xml',
                 'https://raw.githubusercontent.com/rapid7/recog/main/xml/snmp_sysobjid.xml'),
                (self._data_dir / 'manuf.txt',
                 'https://www.wireshark.org/download/automated/data/manuf'),
            ]

            for fpath, url in files:
                self._status_label.setText(f"Downloading {fpath.name}...")
                # Use urlopen with context instead of urlretrieve
                with urllib.request.urlopen(url, context=ssl_context) as response:
                    with open(fpath, 'wb') as f:
                        f.write(response.read())

            self._status_label.setText("Data files downloaded successfully")

        except Exception as e:
            self._status_label.setText(f"Download failed: {e}")

        self._progress_bar.setVisible(False)

    def closeEvent(self, event):
        """Handle dialog close - cancel any running poll"""
        if self._worker and self._worker.isRunning():
            self._worker.cancel()
            # Wait for thread to finish before allowing close
            if not self._worker.wait_for_finish(3000):
                # Force quit if it doesn't stop
                self._worker.terminate()
                self._worker.wait(1000)
        event.accept()

    def _apply_theme(self, theme: ThemeColors):
        self.setStyleSheet(f"""
            QDialog {{
                background-color: {theme.bg_primary};
                color: {theme.text_primary};
            }}
            QGroupBox {{
                font-weight: bold;
                border: 1px solid {theme.border_dim};
                border-radius: 6px;
                margin-top: 12px;
                padding-top: 10px;
                background-color: {theme.bg_secondary};
            }}
            QGroupBox::title {{
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px;
                color: {theme.accent};
            }}
            QLineEdit, QComboBox {{
                background-color: {theme.bg_tertiary};
                border: 1px solid {theme.border_dim};
                border-radius: 4px;
                padding: 6px;
                color: {theme.text_primary};
            }}
            QLineEdit:focus, QComboBox:focus {{
                border-color: {theme.accent};
            }}
            QLineEdit:disabled, QComboBox:disabled {{
                background-color: {theme.bg_primary};
                color: {theme.text_secondary};
            }}
            QTextEdit {{
                background-color: {theme.bg_secondary};
                border: 1px solid {theme.border_dim};
                border-radius: 4px;
                color: {theme.text_primary};
            }}
            QTableWidget {{
                background-color: {theme.bg_secondary};
                border: 1px solid {theme.border_dim};
                gridline-color: {theme.border_dim};
                color: {theme.text_primary};
            }}
            QTableWidget::item {{
                padding: 4px;
            }}
            QHeaderView::section {{
                background-color: {theme.bg_tertiary};
                color: {theme.text_primary};
                padding: 6px;
                border: none;
                border-bottom: 1px solid {theme.border_dim};
            }}
            QPushButton {{
                background-color: {theme.bg_tertiary};
                border: 1px solid {theme.border_dim};
                border-radius: 4px;
                padding: 8px 16px;
                color: {theme.text_primary};
            }}
            QPushButton:hover {{
                background-color: {theme.accent};
                color: {theme.bg_primary};
            }}
            QPushButton:disabled {{
                opacity: 0.5;
            }}
            QTabWidget::pane {{
                border: 1px solid {theme.border_dim};
                background-color: {theme.bg_secondary};
            }}
            QTabBar::tab {{
                background-color: {theme.bg_tertiary};
                color: {theme.text_secondary};
                padding: 8px 16px;
                border: 1px solid {theme.border_dim};
                border-bottom: none;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }}
            QTabBar::tab:selected {{
                background-color: {theme.bg_secondary};
                color: {theme.accent};
            }}
            QLabel {{
                color: {theme.text_primary};
            }}
            QProgressBar {{
                border: 1px solid {theme.border_dim};
                border-radius: 4px;
                background-color: {theme.bg_tertiary};
            }}
            QProgressBar::chunk {{
                background-color: {theme.accent};
            }}
            QCheckBox {{
                color: {theme.text_primary};
            }}
            QCheckBox::indicator {{
                width: 16px;
                height: 16px;
                border: 1px solid {theme.border_dim};
                border-radius: 3px;
                background-color: {theme.bg_tertiary};
            }}
            QCheckBox::indicator:checked {{
                background-color: {theme.accent};
            }}
            QScrollArea {{
                background-color: transparent;
                border: none;
            }}
            QScrollBar:vertical {{
                background-color: {theme.bg_secondary};
                width: 10px;
                border-radius: 5px;
            }}
            QScrollBar::handle:vertical {{
                background-color: {theme.border_dim};
                border-radius: 5px;
                min-height: 20px;
            }}
            QScrollBar::handle:vertical:hover {{
                background-color: {theme.accent};
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height: 0px;
            }}
        """)