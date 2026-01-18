#!/usr/bin/env python3
"""
Secure Cartography - Security Widget
=====================================

PyQt6 widget for CVE vulnerability analysis of discovered network devices.
Reads CSV export from Secure Cartography, maps platforms to CPE format,
and queries NIST NVD for known vulnerabilities.

Standalone testing:
    python security_widget.py

Integration:
    from .widgets.security_widget import SecurityWidget
    widget = SecurityWidget(theme_manager=theme_manager)
"""

import sys
import re
import csv
import json
import sqlite3
import time
import logging
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Tuple, TYPE_CHECKING
from datetime import datetime, timezone

from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QSplitter,
    QTableView, QHeaderView, QProgressBar, QPushButton, QLabel,
    QFileDialog, QComboBox, QLineEdit, QTextEdit, QGroupBox,
    QMessageBox, QStyledItemDelegate, QStyle, QTabWidget,
    QDialog, QFormLayout, QDialogButtonBox, QSpinBox, QCheckBox,
    QFrame, QScrollArea
)
from PyQt6.QtCore import (
    Qt, QThread, pyqtSignal, QAbstractTableModel, QModelIndex,
    QSortFilterProxyModel, QSettings
)
from PyQt6.QtGui import QColor, QBrush, QFont, QIcon

import requests
from urllib.parse import quote

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ============================================================================
# Platform Parser - Extensible vendor/version extraction
# ============================================================================

@dataclass
class ParsedPlatform:
    """Result of parsing a platform string"""
    raw: str
    vendor: str
    product: str
    version: str
    cpe_vendor: str
    cpe_product: str
    cpe_version: str
    confidence: str  # high, medium, low, manual
    device_count: int = 1
    device_names: list = field(default_factory=list)  # Track device hostnames/IPs

    def to_cpe(self) -> str:
        """Generate CPE 2.3 string"""
        if not all([self.cpe_vendor, self.cpe_product, self.cpe_version]):
            return ""
        return f"cpe:2.3:o:{self.cpe_vendor}:{self.cpe_product}:{self.cpe_version}:*:*:*:*:*:*:*"


class PlatformParser:
    """
    Extensible parser for network device platform strings.

    Extracts vendor, product, and version from strings like:
        "Juniper JUNOS 14.1X53-D40.8"
        "Cisco IOS 12.2(54)SG1"
        "Arista EOS 4.23.3M"

    Users can add custom patterns via add_pattern() or patterns.json
    """

    # Default patterns: (regex, vendor, product, cpe_vendor, cpe_product, version_group)
    # version_group is the regex group index containing the version
    DEFAULT_PATTERNS = [
        # Juniper - multiple formats
        (r"Juniper\s+JUNOS?\s+(?:OS\s+)?(\d+\.\d+[A-Z0-9\-\.]+)",
         "Juniper", "JUNOS", "juniper", "junos", 1),
        (r"JUNOS\s+(\d+\.\d+[A-Z0-9\-\.]+)",
         "Juniper", "JUNOS", "juniper", "junos", 1),

        # Arista EOS - including vEOS-lab and other variants
        (r"Arista\s+(?:vEOS[^\s]*\s+)?EOS\s+(\d+\.\d+\.\d+[A-Z]*)",
         "Arista", "EOS", "arista", "eos", 1),
        (r"Arista\s+Networks?\s+(?:vEOS[^\s]*\s+)?EOS\s+(\d+\.\d+\.\d+[A-Z]*)",
         "Arista", "EOS", "arista", "eos", 1),
        (r"vEOS[^\s]*\s+EOS\s+(\d+\.\d+\.\d+[A-Z]*)",
         "Arista", "EOS", "arista", "eos", 1),

        # Cisco IOS - flexible patterns for various formats
        # Handles: "Cisco IOS IOS 15.2(4.0.55)E", "Cisco 7200 IOS 15.2(4)M11", "Cisco IOSv IOS 15.6(2)T"
        (r"Cisco\s+(?:IOS[v]?\s+)?(?:IOS\s+)?(\d+\.\d+\([0-9\.]+\)[A-Za-z0-9]*)",
         "Cisco", "IOS", "cisco", "ios", 1),
        (r"Cisco\s+\d+\s+IOS\s+(\d+\.\d+\([0-9\.]+\)[A-Za-z0-9]*)",
         "Cisco", "IOS", "cisco", "ios", 1),
        (r"Cisco\s+IOSv\s+IOS\s+(\d+\.\d+\([0-9\.]+\)[A-Za-z0-9]*)",
         "Cisco", "IOS", "cisco", "ios", 1),

        # Cisco IOS-XE / IOS-XR / NX-OS
        (r"Cisco\s+IOS[- ]?XE\s+[Ss]oftware[,\s]+Version\s+(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "Cisco", "IOS-XE", "cisco", "ios_xe", 1),
        (r"Cisco\s+IOS[- ]?XR\s+[Ss]oftware[,\s]+Version\s+(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "Cisco", "IOS-XR", "cisco", "ios_xr", 1),
        (r"Cisco\s+NX-?OS.*Version\s+(\d+\.\d+\([0-9]+\)[A-Za-z0-9\.]*)",
         "Cisco", "NX-OS", "cisco", "nx-os", 1),
        (r"Cisco\s+IOS\s+(?:Software,?\s+)?(?:Version\s+)?(\d+\.\d+\([0-9]+\)[A-Za-z0-9]*)",
         "Cisco", "IOS", "cisco", "ios", 1),

        # Fallback Cisco pattern - very flexible
        (r"[Cc]isco.*IOS.*?(\d+\.\d+\([0-9\.]+\)[A-Za-z0-9]*)",
         "Cisco", "IOS", "cisco", "ios", 1),

        # Palo Alto PAN-OS
        (r"Palo\s+Alto\s+Networks?\s+PAN-?OS\s+(\d+\.\d+\.\d+[A-Za-z0-9\-]*)",
         "Palo Alto", "PAN-OS", "paloaltonetworks", "pan-os", 1),
        (r"PAN-?OS\s+(\d+\.\d+\.\d+[A-Za-z0-9\-]*)",
         "Palo Alto", "PAN-OS", "paloaltonetworks", "pan-os", 1),

        # Fortinet FortiOS
        (r"Fortinet\s+FortiOS\s+[vV]?(\d+\.\d+\.\d+[A-Za-z0-9\-]*)",
         "Fortinet", "FortiOS", "fortinet", "fortios", 1),
        (r"FortiOS\s+[vV]?(\d+\.\d+\.\d+[A-Za-z0-9\-]*)",
         "Fortinet", "FortiOS", "fortinet", "fortios", 1),
        (r"FortiGate.*[vV](\d+\.\d+\.\d+)",
         "Fortinet", "FortiOS", "fortinet", "fortios", 1),

        # F5 BIG-IP
        (r"BIG-?IP.*(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "F5", "BIG-IP", "f5", "big-ip_access_policy_manager", 1),

        # HPE/Aruba
        (r"ArubaOS[- ]?(?:CX)?[- ]?(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "Aruba", "ArubaOS", "arubanetworks", "arubaos", 1),
        (r"HPE?\s+(?:ProCurve|Comware).*Version\s+(\d+\.\d+[A-Za-z0-9\.]*)",
         "HPE", "Comware", "hp", "comware", 1),

        # Dell / Force10
        (r"Dell\s+(?:Networking\s+)?OS[0-9]*\s+(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "Dell", "OS10", "dell", "os10", 1),

        # Brocade / Ruckus
        (r"Brocade.*FOS\s+[vV]?(\d+\.\d+\.\d+[A-Za-z0-9]*)",
         "Brocade", "FOS", "brocade", "fabric_os", 1),

        # Extreme Networks
        (r"Extreme(?:XOS|Ware)?\s+(\d+\.\d+\.\d+[A-Za-z0-9\.]*)",
         "Extreme", "ExtremeXOS", "extremenetworks", "extremexos", 1),

        # MikroTik RouterOS
        (r"(?:MikroTik\s+)?RouterOS\s+(\d+\.\d+[A-Za-z0-9\.]*)",
         "MikroTik", "RouterOS", "mikrotik", "routeros", 1),

        # Ubiquiti
        (r"Ubiquiti.*EdgeOS\s+[vV]?(\d+\.\d+\.\d+)",
         "Ubiquiti", "EdgeOS", "ubiquiti", "edgeos", 1),
        (r"UniFi.*(\d+\.\d+\.\d+)",
         "Ubiquiti", "UniFi", "ubiquiti", "unifi_controller", 1),
    ]

    def __init__(self, custom_patterns_path: Optional[Path] = None):
        self.patterns = list(self.DEFAULT_PATTERNS)
        self.custom_patterns_path = custom_patterns_path or Path.home() / ".scng" / "platform_patterns.json"
        self._load_custom_patterns()

    def _load_custom_patterns(self):
        """Load user-defined patterns from JSON file"""
        if self.custom_patterns_path.exists():
            try:
                with open(self.custom_patterns_path) as f:
                    custom = json.load(f)
                for p in custom:
                    self.patterns.insert(0, (  # Custom patterns take priority
                        p["regex"],
                        p["vendor"],
                        p["product"],
                        p["cpe_vendor"],
                        p["cpe_product"],
                        p.get("version_group", 1)
                    ))
                logger.info(f"Loaded {len(custom)} custom patterns")
            except Exception as e:
                logger.warning(f"Could not load custom patterns: {e}")

    def save_custom_pattern(self, regex: str, vendor: str, product: str,
                            cpe_vendor: str, cpe_product: str):
        """Save a new custom pattern"""
        self.custom_patterns_path.parent.mkdir(parents=True, exist_ok=True)

        existing = []
        if self.custom_patterns_path.exists():
            with open(self.custom_patterns_path) as f:
                existing = json.load(f)

        existing.append({
            "regex": regex,
            "vendor": vendor,
            "product": product,
            "cpe_vendor": cpe_vendor,
            "cpe_product": cpe_product,
            "version_group": 1
        })

        with open(self.custom_patterns_path, "w") as f:
            json.dump(existing, f, indent=2)

        # Add to active patterns
        self.patterns.insert(0, (regex, vendor, product, cpe_vendor, cpe_product, 1))

    def parse(self, platform_string: str) -> ParsedPlatform:
        """
        Parse a platform string and extract vendor/product/version

        Returns ParsedPlatform with confidence level
        """
        if not platform_string:
            return ParsedPlatform(
                raw=platform_string, vendor="", product="", version="",
                cpe_vendor="", cpe_product="", cpe_version="",
                confidence="low"
            )

        for pattern, vendor, product, cpe_vendor, cpe_product, ver_group in self.patterns:
            match = re.search(pattern, platform_string, re.IGNORECASE)
            if match:
                version = match.group(ver_group)
                cpe_version = self._normalize_version(cpe_vendor, version)

                return ParsedPlatform(
                    raw=platform_string,
                    vendor=vendor,
                    product=product,
                    version=version,
                    cpe_vendor=cpe_vendor,
                    cpe_product=cpe_product,
                    cpe_version=cpe_version,
                    confidence="high"
                )

        # No pattern matched - try generic extraction
        return self._generic_parse(platform_string)

    def _generic_parse(self, platform_string: str) -> ParsedPlatform:
        """Fallback parser for unrecognized platforms"""
        # Try to extract any version-like string
        version_match = re.search(r'(\d+\.\d+[\.\d]*[A-Za-z0-9\-\(\)]*)', platform_string)
        version = version_match.group(1) if version_match else ""

        # Try to identify vendor from known keywords
        vendor = ""
        for keyword in ["Cisco", "Juniper", "Arista", "Palo Alto", "Fortinet",
                        "F5", "HPE", "Dell", "Brocade", "Extreme", "MikroTik"]:
            if keyword.lower() in platform_string.lower():
                vendor = keyword
                break

        return ParsedPlatform(
            raw=platform_string,
            vendor=vendor,
            product="",
            version=version,
            cpe_vendor="",
            cpe_product="",
            cpe_version="",
            confidence="low"
        )

    def _normalize_version(self, vendor: str, version: str) -> str:
        """Normalize version string for CPE format"""
        v = version.lower().strip()

        # Cisco IOS uses parentheses which need escaping in CPE
        if vendor == "cisco":
            v = v.replace("(", "\\(").replace(")", "\\)")

        return v


# ============================================================================
# CVE Cache (simplified from nvd_cache.py for widget integration)
# ============================================================================

SCHEMA = """
CREATE TABLE IF NOT EXISTS cve_records (
    cve_id TEXT PRIMARY KEY,
    description TEXT,
    severity TEXT,
    cvss_v3_score REAL,
    cvss_v3_vector TEXT,
    cvss_v2_score REAL,
    cvss_v2_vector TEXT,
    published_date TEXT,
    last_modified TEXT,
    cached_at TEXT
);

CREATE TABLE IF NOT EXISTS cpe_cve_mapping (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    cve_id TEXT REFERENCES cve_records(cve_id),
    cpe_vendor TEXT,
    cpe_product TEXT,
    version_exact TEXT,
    UNIQUE(cve_id, cpe_vendor, cpe_product, version_exact)
);

CREATE TABLE IF NOT EXISTS synced_versions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    vendor TEXT,
    product TEXT,
    version TEXT,
    cve_count INTEGER DEFAULT 0,
    critical_count INTEGER DEFAULT 0,
    high_count INTEGER DEFAULT 0,
    medium_count INTEGER DEFAULT 0,
    low_count INTEGER DEFAULT 0,
    last_synced TEXT,
    UNIQUE(vendor, product, version)
);

CREATE INDEX IF NOT EXISTS idx_cve_severity ON cve_records(severity);
CREATE INDEX IF NOT EXISTS idx_cpe_mapping ON cpe_cve_mapping(cpe_vendor, cpe_product, version_exact);
"""


class CVECache:
    """SQLite cache for NVD CVE data"""

    NVD_BASE_URL = "https://services.nvd.nist.gov/rest/json/cves/2.0"

    def __init__(self, db_path: Path):
        self.db_path = db_path
        db_path.parent.mkdir(parents=True, exist_ok=True)
        self.conn = sqlite3.connect(str(db_path))
        self.conn.row_factory = sqlite3.Row
        self.conn.executescript(SCHEMA)
        self.conn.commit()

    def close(self):
        self.conn.close()

    def reconnect(self):
        """Reconnect to database - use after external writes"""
        self.conn.close()
        self.conn = sqlite3.connect(str(self.db_path))
        self.conn.row_factory = sqlite3.Row

    def is_version_synced(self, vendor: str, product: str, version: str) -> bool:
        """Check if a version has been synced"""
        row = self.conn.execute(
            "SELECT 1 FROM synced_versions WHERE vendor=? AND product=? AND version=?",
            (vendor.lower(), product.lower(), version.lower())
        ).fetchone()
        return row is not None

    def sync_version(self, vendor: str, product: str, version: str,
                     api_key: Optional[str] = None) -> Dict:
        """
        Sync CVEs for a single version from NVD

        Returns dict with sync results
        """
        cpe = f"cpe:2.3:o:{vendor}:{product}:{version}:*:*:*:*:*:*:*"
        encoded_cpe = quote(cpe, safe='')
        url = f"{self.NVD_BASE_URL}?cpeName={encoded_cpe}"

        headers = {'User-Agent': 'SecureCartography/1.0'}
        if api_key:
            headers['apiKey'] = api_key

        try:
            response = requests.get(url, headers=headers, timeout=30)

            if response.status_code != 200:
                return {"error": f"HTTP {response.status_code}", "cve_count": 0}

            data = response.json()
            vulnerabilities = data.get("vulnerabilities", [])

            severity_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}

            for vuln in vulnerabilities:
                cve_data = vuln.get("cve", {})
                cve_id = cve_data.get("id", "")

                # Extract CVSS data
                metrics = cve_data.get("metrics", {})
                severity, cvss_v3_score, cvss_v3_vector = self._extract_cvss_v3(metrics)
                cvss_v2_score, cvss_v2_vector = self._extract_cvss_v2(metrics)

                if severity in severity_counts:
                    severity_counts[severity] += 1

                # Get description
                description = ""
                for desc in cve_data.get("descriptions", []):
                    if desc.get("lang") == "en":
                        description = desc.get("value", "")
                        break

                # Upsert CVE record
                self.conn.execute("""
                    INSERT INTO cve_records (cve_id, description, severity, 
                        cvss_v3_score, cvss_v3_vector, cvss_v2_score, cvss_v2_vector,
                        published_date, last_modified, cached_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(cve_id) DO UPDATE SET
                        description=excluded.description,
                        severity=excluded.severity,
                        cvss_v3_score=excluded.cvss_v3_score,
                        cached_at=excluded.cached_at
                """, (
                    cve_id, description, severity,
                    cvss_v3_score, cvss_v3_vector, cvss_v2_score, cvss_v2_vector,
                    cve_data.get("published"), cve_data.get("lastModified"),
                    datetime.now(timezone.utc).isoformat()
                ))

                # Add CPE mapping
                self.conn.execute("""
                    INSERT OR IGNORE INTO cpe_cve_mapping 
                    (cve_id, cpe_vendor, cpe_product, version_exact)
                    VALUES (?, ?, ?, ?)
                """, (cve_id, vendor.lower(), product.lower(), version.lower()))

            # Update synced_versions
            self.conn.execute("""
                INSERT INTO synced_versions (vendor, product, version, cve_count,
                    critical_count, high_count, medium_count, low_count, last_synced)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(vendor, product, version) DO UPDATE SET
                    cve_count=excluded.cve_count,
                    critical_count=excluded.critical_count,
                    high_count=excluded.high_count,
                    medium_count=excluded.medium_count,
                    low_count=excluded.low_count,
                    last_synced=excluded.last_synced
            """, (
                vendor.lower(), product.lower(), version.lower(),
                len(vulnerabilities),
                severity_counts["CRITICAL"], severity_counts["HIGH"],
                severity_counts["MEDIUM"], severity_counts["LOW"],
                datetime.now(timezone.utc).isoformat()
            ))

            self.conn.commit()

            return {
                "cve_count": len(vulnerabilities),
                "critical": severity_counts["CRITICAL"],
                "high": severity_counts["HIGH"],
                "medium": severity_counts["MEDIUM"],
                "low": severity_counts["LOW"],
            }

        except Exception as e:
            logger.error(f"Sync failed for {vendor}:{product}:{version}: {e}")
            return {"error": str(e), "cve_count": 0}

    def _extract_cvss_v3(self, metrics: Dict) -> Tuple[str, Optional[float], Optional[str]]:
        """Extract CVSS v3.x severity, score, vector"""
        for key in ["cvssMetricV31", "cvssMetricV30"]:
            if key in metrics and metrics[key]:
                m = metrics[key][0]
                cvss = m.get("cvssData", {})
                return (
                    cvss.get("baseSeverity", "UNKNOWN"),
                    cvss.get("baseScore"),
                    cvss.get("vectorString")
                )
        return ("UNKNOWN", None, None)

    def _extract_cvss_v2(self, metrics: Dict) -> Tuple[Optional[float], Optional[str]]:
        """Extract CVSS v2 score and vector"""
        if "cvssMetricV2" in metrics and metrics["cvssMetricV2"]:
            m = metrics["cvssMetricV2"][0]
            cvss = m.get("cvssData", {})
            return (cvss.get("baseScore"), cvss.get("vectorString"))
        return (None, None)

    def get_cves_for_version(self, vendor: str, product: str, version: str) -> List[Dict]:
        """Get cached CVEs for a version"""
        rows = self.conn.execute("""
            SELECT c.* FROM cve_records c
            JOIN cpe_cve_mapping m ON c.cve_id = m.cve_id
            WHERE m.cpe_vendor = ? AND m.cpe_product = ? AND m.version_exact = ?
            ORDER BY c.cvss_v3_score DESC NULLS LAST
        """, (vendor.lower(), product.lower(), version.lower())).fetchall()
        return [dict(row) for row in rows]

    def get_version_summary(self) -> List[Dict]:
        """Get summary of all synced versions"""
        rows = self.conn.execute("""
            SELECT * FROM synced_versions ORDER BY cve_count DESC
        """).fetchall()
        return [dict(row) for row in rows]

    def get_overall_summary(self) -> Dict:
        """Get overall CVE statistics"""
        row = self.conn.execute("""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN severity='CRITICAL' THEN 1 ELSE 0 END) as critical,
                SUM(CASE WHEN severity='HIGH' THEN 1 ELSE 0 END) as high,
                SUM(CASE WHEN severity='MEDIUM' THEN 1 ELSE 0 END) as medium,
                SUM(CASE WHEN severity='LOW' THEN 1 ELSE 0 END) as low
            FROM cve_records
        """).fetchone()
        return dict(row) if row else {}


# ============================================================================
# Qt Models
# ============================================================================

class PlatformTableModel(QAbstractTableModel):
    """Model for the platform/CPE mapping table"""

    COLUMNS = ["Platform", "Vendor", "Product", "Version", "CPE Version",
               "Devices", "Confidence", "Status"]

    def __init__(self):
        super().__init__()
        self.platforms: List[ParsedPlatform] = []
        self.sync_status: Dict[str, str] = {}  # raw -> status

    def load_platforms(self, platforms: List[ParsedPlatform]):
        self.beginResetModel()
        self.platforms = platforms
        self.sync_status = {p.raw: "pending" for p in platforms}
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.platforms)

    def columnCount(self, parent=QModelIndex()):
        return len(self.COLUMNS)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.COLUMNS[section]
        return None

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self.platforms):
            return None

        p = self.platforms[index.row()]
        col = index.column()

        if role == Qt.ItemDataRole.DisplayRole:
            if col == 0: return p.raw[:60] + "..." if len(p.raw) > 60 else p.raw
            if col == 1: return p.cpe_vendor
            if col == 2: return p.cpe_product
            if col == 3: return p.version
            if col == 4: return p.cpe_version
            if col == 5: return str(p.device_count)
            if col == 6: return p.confidence
            if col == 7: return self.sync_status.get(p.raw, "pending")

        elif role == Qt.ItemDataRole.BackgroundRole:
            status = self.sync_status.get(p.raw, "pending")
            if status == "synced":
                return QBrush(QColor("#d4edda"))
            elif status == "error":
                return QBrush(QColor("#f8d7da"))
            elif status == "syncing":
                return QBrush(QColor("#fff3cd"))
            elif p.confidence == "low":
                return QBrush(QColor("#fff3cd"))

        elif role == Qt.ItemDataRole.ToolTipRole:
            if col == 0:
                return p.raw
            elif col == 4:
                return p.to_cpe()

        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid() or role != Qt.ItemDataRole.EditRole:
            return False

        p = self.platforms[index.row()]
        col = index.column()

        # Allow editing CPE fields
        if col == 1:
            p.cpe_vendor = value.lower().strip()
        elif col == 2:
            p.cpe_product = value.lower().strip()
        elif col == 4:
            p.cpe_version = value.lower().strip()
        else:
            return False

        p.confidence = "manual"
        self.dataChanged.emit(index, index)
        return True

    def flags(self, index):
        flags = super().flags(index)
        # Make CPE columns editable
        if index.column() in [1, 2, 4]:
            flags |= Qt.ItemFlag.ItemIsEditable
        return flags

    def update_status(self, raw: str, status: str):
        """Update sync status for a platform"""
        self.sync_status[raw] = status
        for i, p in enumerate(self.platforms):
            if p.raw == raw:
                idx = self.index(i, 7)
                self.dataChanged.emit(idx, idx)
                break

    def get_selected_for_sync(self, rows: List[int]) -> List[ParsedPlatform]:
        """Get platforms ready for sync (have valid CPE data)"""
        result = []
        for row in rows:
            if row < len(self.platforms):
                p = self.platforms[row]
                if p.cpe_vendor and p.cpe_product and p.cpe_version:
                    result.append(p)
        return result


class SeverityDelegate(QStyledItemDelegate):
    """Custom delegate to paint severity colors that override stylesheets"""

    SEVERITY_COLORS = {
        "CRITICAL": ("#dc2626", "#ffffff"),  # bg, fg
        "HIGH": ("#ea580c", "#ffffff"),
        "MEDIUM": ("#ca8a04", "#ffffff"),
        "LOW": ("#16a34a", "#ffffff"),
    }

    def __init__(self, model, parent=None):
        super().__init__(parent)
        self.model = model

    def paint(self, painter, option, index):
        # Get severity from model
        if index.row() < len(self.model.cves):
            severity = self.model.cves[index.row()].get("severity", "").upper()

            if severity in self.SEVERITY_COLORS:
                bg_color, fg_color = self.SEVERITY_COLORS[severity]

                # Fill background
                painter.fillRect(option.rect, QColor(bg_color))

                # Handle selection highlight
                if option.state & QStyle.StateFlag.State_Selected:
                    painter.fillRect(option.rect, QColor(255, 255, 255, 60))

                # Draw text
                painter.setPen(QColor(fg_color))
                text = index.data(Qt.ItemDataRole.DisplayRole)
                if text:
                    text_rect = option.rect.adjusted(6, 0, -6, 0)
                    painter.drawText(text_rect, Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft, str(text))
                return

        # Fall back to default painting
        super().paint(painter, option, index)


class CVETableModel(QAbstractTableModel):
    """Model for displaying CVE results"""

    COLUMNS = ["CVE ID", "Severity", "CVSS", "Published", "Description"]

    def __init__(self):
        super().__init__()
        self.cves: List[Dict] = []

    def load_cves(self, cves: List[Dict]):
        self.beginResetModel()
        self.cves = cves
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.cves)

    def columnCount(self, parent=QModelIndex()):
        return len(self.COLUMNS)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.COLUMNS[section]
        return None

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self.cves):
            return None

        cve = self.cves[index.row()]
        col = index.column()

        if role == Qt.ItemDataRole.DisplayRole:
            if col == 0: return cve.get("cve_id", "")
            if col == 1: return cve.get("severity", "")
            if col == 2: return str(cve.get("cvss_v3_score", "")) if cve.get("cvss_v3_score") else ""
            if col == 3: return cve.get("published_date", "")[:10] if cve.get("published_date") else ""
            if col == 4:
                desc = cve.get("description", "")
                return desc[:100] + "..." if len(desc) > 100 else desc

        elif role == Qt.ItemDataRole.BackgroundRole:
            severity = cve.get("severity", "").upper()
            # Severity background colors with good contrast for white/light text
            colors = {
                "CRITICAL": QColor("#dc2626"),  # Red-600
                "HIGH": QColor("#ea580c"),  # Orange-600
                "MEDIUM": QColor("#ca8a04"),  # Yellow-600
                "LOW": QColor("#16a34a"),  # Green-600
            }
            if severity in colors:
                return QBrush(colors[severity])

        elif role == Qt.ItemDataRole.ToolTipRole:
            if col == 4:
                return cve.get("description", "")

        return None


class CachedVersionsModel(QAbstractTableModel):
    """Model for displaying cached/synced versions from the database"""

    COLUMNS = ["Vendor", "Product", "Version", "CVEs", "Critical", "High", "Medium", "Low", "Last Synced"]

    def __init__(self):
        super().__init__()
        self.versions: List[Dict] = []

    def load_versions(self, versions: List[Dict]):
        self.beginResetModel()
        self.versions = versions
        self.endResetModel()

    def rowCount(self, parent=QModelIndex()):
        return len(self.versions)

    def columnCount(self, parent=QModelIndex()):
        return len(self.COLUMNS)

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Horizontal and role == Qt.ItemDataRole.DisplayRole:
            return self.COLUMNS[section]
        return None

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid() or index.row() >= len(self.versions):
            return None

        v = self.versions[index.row()]
        col = index.column()

        if role == Qt.ItemDataRole.DisplayRole:
            if col == 0: return v.get("vendor", "")
            if col == 1: return v.get("product", "")
            if col == 2: return v.get("version", "")
            if col == 3: return str(v.get("cve_count", 0))
            if col == 4: return str(v.get("critical_count", 0))
            if col == 5: return str(v.get("high_count", 0))
            if col == 6: return str(v.get("medium_count", 0))
            if col == 7: return str(v.get("low_count", 0))
            if col == 8:
                synced = v.get("last_synced", "")
                return synced[:10] if synced else ""

        elif role == Qt.ItemDataRole.BackgroundRole:
            # Highlight rows with critical CVEs
            critical = v.get("critical_count", 0)
            high = v.get("high_count", 0)
            if critical > 0:
                return QBrush(QColor("#f8d7da"))
            elif high > 5:
                return QBrush(QColor("#ffe5d0"))

        elif role == Qt.ItemDataRole.TextAlignmentRole:
            if col >= 3:  # Numeric columns
                return Qt.AlignmentFlag.AlignCenter

        return None

    def get_version_at(self, row: int) -> Optional[Dict]:
        """Get version data at row"""
        if 0 <= row < len(self.versions):
            return self.versions[row]
        return None


# ============================================================================
# Sync Worker Thread
# ============================================================================

class SyncWorker(QThread):
    """Background worker for NVD sync operations"""

    progress = pyqtSignal(int, int, str)  # current, total, message
    version_complete = pyqtSignal(str, dict)  # raw_platform, result
    finished_all = pyqtSignal(dict)  # summary

    def __init__(self, db_path: Path, platforms: List[ParsedPlatform],
                 api_key: Optional[str] = None, delay: float = 6.0):
        super().__init__()
        self.db_path = db_path  # Store path, create connection in run()
        self.platforms = platforms
        self.api_key = api_key
        self.delay = delay if not api_key else 0.6
        self._stop = False

    def stop(self):
        self._stop = True

    def run(self):
        # Create cache connection in this thread
        cache = CVECache(self.db_path)

        total = len(self.platforms)
        results = {"synced": 0, "errors": 0, "total_cves": 0}

        for i, p in enumerate(self.platforms):
            if self._stop:
                break

            self.progress.emit(i + 1, total, f"Syncing {p.cpe_vendor}:{p.cpe_product}:{p.cpe_version}")

            result = cache.sync_version(p.cpe_vendor, p.cpe_product, p.cpe_version, self.api_key)

            if "error" in result:
                results["errors"] += 1
                self.version_complete.emit(p.raw, {"status": "error", **result})
            else:
                results["synced"] += 1
                results["total_cves"] += result.get("cve_count", 0)
                self.version_complete.emit(p.raw, {"status": "synced", **result})

            # Rate limiting delay
            if i < total - 1 and not self._stop:
                time.sleep(self.delay)

        cache.close()
        self.finished_all.emit(results)


# ============================================================================
# Main Widget
# ============================================================================

class CVEDetailDialog(QDialog):
    """Dialog showing full CVE details for a version"""

    def __init__(self, vendor: str, product: str, version: str,
                 cves: List[Dict], parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"CVEs for {vendor}:{product}:{version}")
        self.setMinimumSize(900, 600)
        self.cves = cves

        layout = QVBoxLayout(self)

        # Summary header
        header = QLabel(f"<b>{len(cves)} CVEs</b> affecting {vendor} {product} {version}")
        header.setStyleSheet("font-size: 14px; padding: 8px;")
        layout.addWidget(header)

        # Severity summary
        severity_counts = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0, "LOW": 0}
        for cve in cves:
            sev = cve.get("severity", "").upper()
            if sev in severity_counts:
                severity_counts[sev] += 1

        severity_text = " | ".join([
            f"<span style='color: #dc3545;'>Critical: {severity_counts['CRITICAL']}</span>",
            f"<span style='color: #fd7e14;'>High: {severity_counts['HIGH']}</span>",
            f"<span style='color: #ffc107;'>Medium: {severity_counts['MEDIUM']}</span>",
            f"<span style='color: #28a745;'>Low: {severity_counts['LOW']}</span>",
        ])
        severity_label = QLabel(severity_text)
        severity_label.setStyleSheet("padding: 4px 8px;")
        layout.addWidget(severity_label)

        # CVE table
        self.cve_model = CVETableModel()
        self.cve_model.load_cves(cves)

        self.cve_table = QTableView()
        self.cve_table.setModel(self.cve_model)
        self.cve_table.setItemDelegate(SeverityDelegate(self.cve_model, self.cve_table))
        self.cve_table.setAlternatingRowColors(False)  # Delegate handles colors
        self.cve_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.cve_table.horizontalHeader().setStretchLastSection(True)
        self.cve_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.cve_table.doubleClicked.connect(self._show_cve_detail)
        layout.addWidget(self.cve_table)

        # Detail panel for selected CVE
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setMaximumHeight(150)
        self.detail_text.setPlaceholderText("Select a CVE to view details...")
        layout.addWidget(self.detail_text)

        self.cve_table.selectionModel().selectionChanged.connect(self._on_selection_changed)

        # Buttons
        btn_layout = QHBoxLayout()

        export_btn = QPushButton("Export CSV")
        export_btn.clicked.connect(self._export_csv)
        btn_layout.addWidget(export_btn)

        btn_layout.addStretch()

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)

        layout.addLayout(btn_layout)

    def _on_selection_changed(self):
        """Show detail for selected CVE"""
        indexes = self.cve_table.selectedIndexes()
        if not indexes:
            return

        row = indexes[0].row()
        if row < len(self.cves):
            cve = self.cves[row]
            detail = f"<b>{cve.get('cve_id', '')}</b><br><br>"
            detail += f"<b>Severity:</b> {cve.get('severity', 'N/A')} "
            detail += f"(CVSS: {cve.get('cvss_v3_score', 'N/A')})<br>"
            detail += f"<b>Published:</b> {cve.get('published_date', 'N/A')[:10] if cve.get('published_date') else 'N/A'}<br><br>"
            detail += f"<b>Description:</b><br>{cve.get('description', 'No description available')}"
            self.detail_text.setHtml(detail)

    def _show_cve_detail(self, index):
        """Open CVE in browser on double-click"""
        row = index.row()
        if row < len(self.cves):
            cve_id = self.cves[row].get('cve_id', '')
            if cve_id:
                import webbrowser
                webbrowser.open(f"https://nvd.nist.gov/vuln/detail/{cve_id}")

    def _export_csv(self):
        """Export CVE list to CSV"""
        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export CVEs", f"cves_export.csv", "CSV Files (*.csv)"
        )
        if not filepath:
            return

        try:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(['CVE ID', 'Severity', 'CVSS v3', 'Published', 'Description'])
                for cve in self.cves:
                    writer.writerow([
                        cve.get('cve_id', ''),
                        cve.get('severity', ''),
                        cve.get('cvss_v3_score', ''),
                        cve.get('published_date', '')[:10] if cve.get('published_date') else '',
                        cve.get('description', '')
                    ])
            QMessageBox.information(self, "Export Complete", f"Exported {len(self.cves)} CVEs to {filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))


# ============================================================================
# Help Dialog
# ============================================================================

class SecurityHelpDialog(QDialog):
    """Help dialog explaining the security analysis workflow"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Security Analysis - Help")
        self.setMinimumSize(700, 500)

        layout = QVBoxLayout(self)

        # Header
        header = QLabel("🔒 Security Analysis Help")
        header.setStyleSheet("font-size: 18px; font-weight: bold; padding: 8px;")
        layout.addWidget(header)

        # Help content
        help_text = QTextEdit()
        help_text.setReadOnly(True)
        help_text.setHtml(self._get_help_content())
        layout.addWidget(help_text)

        # Close button
        close_btn = QPushButton("Close")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignmentFlag.AlignRight)

    def _get_help_content(self) -> str:
        return """
        <style>
            h2 { color: #2563eb; margin-top: 16px; }
            h3 { color: #64748b; margin-top: 12px; }
            code { background: #f1f5f9; padding: 2px 6px; border-radius: 3px; }
            .note { background: #fef3c7; padding: 8px; border-radius: 4px; margin: 8px 0; }
        </style>

        <h2>Overview</h2>
        <p>The Security Analysis widget identifies known vulnerabilities (CVEs) affecting 
        your discovered network devices by querying the <b>NIST National Vulnerability Database (NVD)</b>.</p>

        <h2>Workflow</h2>
        <h3>1. Export Discovery Results</h3>
        <p>From Secure Cartography's Map Viewer, use <b>Export CSV</b> to save your discovered devices. 
        The CSV must include a <code>platform</code> column with OS/version strings.</p>

        <h3>2. Load CSV</h3>
        <p>Click <b>Load CSV</b> to import the discovery export. The widget will:</p>
        <ul>
            <li>Parse platform strings (e.g., "Cisco IOS 15.2(4)M11")</li>
            <li>Map to CPE format (e.g., cisco:ios:15.2(4)m11)</li>
            <li>Show confidence level for each mapping</li>
        </ul>

        <h3>3. Review Mappings</h3>
        <p>Check the <b>Discovered Platforms</b> table:</p>
        <ul>
            <li><b>High confidence</b> - Auto-mapped, ready to sync</li>
            <li><b>Low confidence</b> - May need manual correction</li>
            <li>Double-click Vendor/Product/CPE Version columns to edit</li>
        </ul>

        <h3>4. Sync with NVD</h3>
        <p>Select rows and click <b>Sync Selected</b> to query NIST NVD for vulnerabilities.</p>
        <div class="note">
            <b>Rate Limiting:</b> NVD allows 5 requests per 30 seconds without an API key.
            For faster syncing, get a free API key from <a href="https://nvd.nist.gov/developers/request-an-api-key">nvd.nist.gov</a>
        </div>

        <h3>5. Review Results</h3>
        <ul>
            <li><b>Cached Versions</b> tab - All synced versions with CVE counts</li>
            <li><b>CVEs</b> tab - Detailed CVE list for selected version</li>
            <li>Double-click a CVE to open in NVD website</li>
            <li>Double-click a cached version for detailed CVE dialog</li>
        </ul>

        <h2>Data Sources</h2>
        <table border="0" cellpadding="4">
            <tr><td><b>NVD</b></td><td>NIST National Vulnerability Database - authoritative CVE source</td></tr>
            <tr><td><b>CPE</b></td><td>Common Platform Enumeration - standardized naming for platforms</td></tr>
            <tr><td><b>CVSS</b></td><td>Common Vulnerability Scoring System - severity ratings</td></tr>
        </table>

        <h2>Cache Location</h2>
        <p>CVE data is cached locally at: <code>~/.scng/cve_cache.db</code></p>
        <p>Custom platform patterns: <code>~/.scng/platform_patterns.json</code></p>

        <h2>Tips</h2>
        <ul>
            <li>Use <b>View Cache</b> to see previously synced data without loading a new CSV</li>
            <li>Check <b>Force Re-sync</b> to refresh CVE data for already-cached versions</li>
            <li>Use <b>Add Pattern</b> to teach the parser new platform formats</li>
            <li>Export full reports via <b>Export Report</b> button</li>
        </ul>
        """


class SecurityWidget(QWidget):
    """
    Main security analysis widget for Secure Cartography

    Features:
    - Load CSV export from discovery
    - Parse platforms and map to CPE format
    - Edit CPE mappings for unrecognized platforms
    - Sync with NIST NVD for vulnerability data
    - View CVE details by version
    - Load and view cached data without new CSV
    - Export vulnerability reports
    - Theme integration with SC
    """

    def __init__(self, db_path: Optional[Path] = None, theme_manager=None, parent=None):
        super().__init__(parent)

        self.db_path = db_path or Path.home() / ".scng" / "cve_cache.db"
        self.cache = CVECache(self.db_path)
        self.parser = PlatformParser()
        self.worker: Optional[SyncWorker] = None
        self.current_csv_path: Optional[str] = None
        self.theme_manager = theme_manager
        self._loaded_platforms = []  # Track loaded platforms for export

        self.settings = QSettings("SecureCartography", "Security")

        self._init_ui()

        # Apply theme if provided
        if theme_manager:
            self.apply_theme(theme_manager.theme)

        # Load cached data on startup
        self._load_from_cache()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)

        # Top controls - Row 1
        top_bar = QHBoxLayout()

        self.load_btn = QPushButton("Load CSV")
        self.load_btn.clicked.connect(self._load_csv)
        top_bar.addWidget(self.load_btn)

        self.load_cache_btn = QPushButton("View Cache")
        self.load_cache_btn.setToolTip("View previously synced versions from cache")
        self.load_cache_btn.clicked.connect(self._load_from_cache)
        top_bar.addWidget(self.load_cache_btn)

        top_bar.addSpacing(20)

        top_bar.addWidget(QLabel("API Key:"))
        self.api_key_input = QLineEdit()
        self.api_key_input.setPlaceholderText("Optional - increases rate limit")
        self.api_key_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.api_key_input.setText(self.settings.value("nvd_api_key", ""))
        self.api_key_input.setMaximumWidth(200)
        top_bar.addWidget(self.api_key_input)

        top_bar.addStretch()

        # Help button
        self.help_btn = QPushButton("? Help")
        self.help_btn.setToolTip("Show help for security analysis workflow")
        self.help_btn.clicked.connect(self._show_help)
        top_bar.addWidget(self.help_btn)

        top_bar.addSpacing(10)

        # Force re-sync checkbox
        self.force_sync_cb = QCheckBox("Force Re-sync")
        self.force_sync_cb.setToolTip("Re-sync versions even if already in cache")
        top_bar.addWidget(self.force_sync_cb)

        self.sync_btn = QPushButton("Sync Selected")
        self.sync_btn.clicked.connect(self._start_sync)
        self.sync_btn.setEnabled(False)
        top_bar.addWidget(self.sync_btn)

        self.stop_btn = QPushButton("Stop")
        self.stop_btn.clicked.connect(self._stop_sync)
        self.stop_btn.setEnabled(False)
        top_bar.addWidget(self.stop_btn)

        layout.addLayout(top_bar)

        # Progress bar
        self.progress = QProgressBar()
        self.progress.setVisible(False)
        layout.addWidget(self.progress)

        self.status_label = QLabel("Load a CSV export from Secure Cartography to begin")
        layout.addWidget(self.status_label)

        # Main content splitter
        splitter = QSplitter(Qt.Orientation.Vertical)

        # Platform table
        platform_group = QGroupBox("Discovered Platforms")
        platform_layout = QVBoxLayout(platform_group)

        self.platform_model = PlatformTableModel()
        self.platform_table = QTableView()
        self.platform_table.setModel(self.platform_model)
        self.platform_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.platform_table.setAlternatingRowColors(True)
        self.platform_table.horizontalHeader().setStretchLastSection(True)
        self.platform_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.platform_table.selectionModel().selectionChanged.connect(self._on_platform_selected)

        platform_layout.addWidget(self.platform_table)

        # Platform action buttons
        platform_actions = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All Valid")
        self.select_all_btn.clicked.connect(self._select_all_valid)
        platform_actions.addWidget(self.select_all_btn)

        self.add_pattern_btn = QPushButton("Add Pattern...")
        self.add_pattern_btn.clicked.connect(self._add_custom_pattern)
        platform_actions.addWidget(self.add_pattern_btn)

        platform_actions.addStretch()

        self.platform_count_label = QLabel("")
        platform_actions.addWidget(self.platform_count_label)

        platform_layout.addLayout(platform_actions)
        splitter.addWidget(platform_group)

        # CVE results (tabs)
        results_group = QGroupBox("CVE Results")
        results_layout = QVBoxLayout(results_group)

        self.results_tabs = QTabWidget()

        # Cached versions tab (from database)
        cached_widget = QWidget()
        cached_layout = QVBoxLayout(cached_widget)

        cached_header = QHBoxLayout()
        self.cached_count_label = QLabel("0 versions in cache")
        cached_header.addWidget(self.cached_count_label)
        cached_header.addStretch()

        refresh_cache_btn = QPushButton("Refresh")
        refresh_cache_btn.clicked.connect(self._load_from_cache)
        cached_header.addWidget(refresh_cache_btn)

        cached_layout.addLayout(cached_header)

        self.cached_model = CachedVersionsModel()
        self.cached_table = QTableView()
        self.cached_table.setModel(self.cached_model)
        self.cached_table.setAlternatingRowColors(True)
        self.cached_table.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.cached_table.horizontalHeader().setStretchLastSection(True)
        self.cached_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.cached_table.doubleClicked.connect(self._on_cached_version_double_click)
        cached_layout.addWidget(self.cached_table)
        self.results_tabs.addTab(cached_widget, "Cached Versions")

        # Summary tab
        summary_widget = QWidget()
        summary_layout = QVBoxLayout(summary_widget)
        self.summary_text = QTextEdit()
        self.summary_text.setReadOnly(True)
        summary_layout.addWidget(self.summary_text)
        self.results_tabs.addTab(summary_widget, "Summary")

        # CVE list tab (for selected version)
        cve_widget = QWidget()
        cve_layout = QVBoxLayout(cve_widget)

        self.cve_version_label = QLabel("Select a version to view CVEs")
        cve_layout.addWidget(self.cve_version_label)

        self.cve_model = CVETableModel()
        self.cve_table = QTableView()
        self.cve_table.setModel(self.cve_model)
        self.cve_table.setItemDelegate(SeverityDelegate(self.cve_model, self.cve_table))
        self.cve_table.setAlternatingRowColors(False)  # Delegate handles colors
        self.cve_table.horizontalHeader().setStretchLastSection(True)
        self.cve_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.cve_table.doubleClicked.connect(self._on_cve_double_click)
        cve_layout.addWidget(self.cve_table)
        self.results_tabs.addTab(cve_widget, "CVEs")

        results_layout.addWidget(self.results_tabs)

        # Export buttons (always visible, below tabs)
        export_row = QHBoxLayout()
        export_row.addStretch()

        export_report_btn = QPushButton("📊 Export Full Report")
        export_report_btn.setToolTip("Export all CVEs with affected platforms to Excel")
        export_report_btn.clicked.connect(self._export_full_report)
        export_row.addWidget(export_report_btn)

        export_devices_btn = QPushButton("📋 Export by Device")
        export_devices_btn.setToolTip("Export vulnerability summary per device to Excel")
        export_devices_btn.clicked.connect(self._export_devices_report)
        export_row.addWidget(export_devices_btn)

        results_layout.addLayout(export_row)

        splitter.addWidget(results_group)

        splitter.setSizes([300, 200])
        layout.addWidget(splitter)

        self._update_summary()

    def _load_csv(self):
        """Load CSV export from Secure Cartography"""
        filepath, _ = QFileDialog.getOpenFileName(
            self, "Load Discovery CSV", "", "CSV Files (*.csv);;All Files (*)"
        )
        if not filepath:
            return

        self.current_csv_path = filepath

        try:
            # Track platform -> list of device names
            platforms_dict = {}  # platform_string -> {'count': int, 'devices': list}

            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Look for platform column (flexible naming)
                    platform = None
                    for key in ['platform', 'Platform', 'os', 'OS', 'software', 'version']:
                        if key in row and row[key]:
                            platform = row[key].strip()
                            break

                    # Get device identifier (hostname or IP)
                    device_name = None
                    for key in ['hostname', 'Hostname', 'name', 'Name', 'ip', 'IP',
                                'ip_address', 'IP Address', 'management_ip', 'device']:
                        if key in row and row[key]:
                            device_name = row[key].strip()
                            break

                    if platform:
                        if platform not in platforms_dict:
                            platforms_dict[platform] = {'count': 0, 'devices': []}
                        platforms_dict[platform]['count'] += 1
                        if device_name and device_name not in platforms_dict[platform]['devices']:
                            platforms_dict[platform]['devices'].append(device_name)

            # Parse each unique platform
            parsed = []
            for platform_str, info in platforms_dict.items():
                p = self.parser.parse(platform_str)
                p.device_count = info['count']
                p.device_names = info['devices']
                parsed.append(p)

            # Sort by device count descending
            parsed.sort(key=lambda x: x.device_count, reverse=True)

            self.platform_model.load_platforms(parsed)

            # Store for export
            self._loaded_platforms = parsed

            self.sync_btn.setEnabled(True)

            valid = sum(1 for p in parsed if p.confidence == "high")
            self.platform_count_label.setText(
                f"{len(parsed)} unique platforms ({valid} auto-mapped)"
            )
            total_devices = sum(info['count'] for info in platforms_dict.values())
            self.status_label.setText(f"Loaded {total_devices} devices from {filepath}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load CSV: {e}")
            logger.exception("CSV load error")

    def _load_from_cache(self):
        """Load and display previously synced versions from cache"""
        # Reconnect to ensure we see latest data
        self.cache.reconnect()

        versions = self.cache.get_version_summary()
        self.cached_model.load_versions(versions)
        self.cached_count_label.setText(f"{len(versions)} versions in cache")

        if versions:
            self.results_tabs.setCurrentIndex(0)  # Switch to cached versions tab
            self.status_label.setText(f"Loaded {len(versions)} versions from cache ({self.db_path})")
        else:
            self.status_label.setText("Cache is empty - load a CSV and sync to populate")

        self._update_summary()

    def _on_cached_version_double_click(self, index):
        """Show CVE detail dialog for double-clicked cached version"""
        v = self.cached_model.get_version_at(index.row())
        if not v:
            return

        cves = self.cache.get_cves_for_version(v['vendor'], v['product'], v['version'])

        dialog = CVEDetailDialog(
            v['vendor'], v['product'], v['version'], cves, self
        )
        dialog.exec()

    def _on_cve_double_click(self, index):
        """Open CVE in browser on double-click"""
        row = index.row()
        if row < len(self.cve_model.cves):
            cve_id = self.cve_model.cves[row].get('cve_id', '')
            if cve_id:
                import webbrowser
                webbrowser.open(f"https://nvd.nist.gov/vuln/detail/{cve_id}")

    def _export_full_report(self):
        """Export full vulnerability report to Excel with affected devices"""
        versions = self.cache.get_version_summary()
        if not versions:
            QMessageBox.warning(self, "No Data", "No cached data to export.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export Vulnerability Report",
            f"vulnerability_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        # Build lookup from loaded platforms (if available)
        platform_devices = {}
        if hasattr(self, '_loaded_platforms') and self._loaded_platforms:
            for p in self._loaded_platforms:
                key = f"{p.cpe_vendor}:{p.cpe_product}:{p.cpe_version}"
                platform_devices[key] = {
                    'count': p.device_count,
                    'devices': p.device_names
                }

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "CVE Report"

            # Severity colors
            severity_fills = {
                'CRITICAL': PatternFill('solid', fgColor='DC2626'),
                'HIGH': PatternFill('solid', fgColor='EA580C'),
                'MEDIUM': PatternFill('solid', fgColor='CA8A04'),
                'LOW': PatternFill('solid', fgColor='16A34A'),
            }
            white_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill('solid', fgColor='1F2937')
            header_font = Font(bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ['Vendor', 'Product', 'Version', 'Device Count', 'Affected Devices',
                       'CVE ID', 'Severity', 'CVSS v3', 'Published', 'Description']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # Data rows
            row_num = 2
            for v in versions:
                cves = self.cache.get_cves_for_version(v['vendor'], v['product'], v['version'])

                # Look up device info
                key = f"{v['vendor']}:{v['product']}:{v['version']}"
                device_info = platform_devices.get(key, {'count': 0, 'devices': []})
                device_count = device_info['count'] or ''
                device_list = '; '.join(device_info['devices'])  # No limit in Excel

                if cves:
                    for cve in cves:
                        severity = cve.get('severity', '').upper()

                        ws.cell(row=row_num, column=1, value=v['vendor'])
                        ws.cell(row=row_num, column=2, value=v['product'])
                        ws.cell(row=row_num, column=3, value=v['version'])
                        ws.cell(row=row_num, column=4, value=device_count)
                        ws.cell(row=row_num, column=5, value=device_list)
                        ws.cell(row=row_num, column=6, value=cve.get('cve_id', ''))

                        sev_cell = ws.cell(row=row_num, column=7, value=severity)
                        if severity in severity_fills:
                            sev_cell.fill = severity_fills[severity]
                            sev_cell.font = white_font
                            sev_cell.alignment = Alignment(horizontal='center')

                        ws.cell(row=row_num, column=8, value=cve.get('cvss_v3_score', ''))
                        ws.cell(row=row_num, column=9,
                                value=cve.get('published_date', '')[:10] if cve.get('published_date') else '')

                        # Full description - no truncation in Excel
                        desc_cell = ws.cell(row=row_num, column=10, value=cve.get('description', ''))
                        desc_cell.alignment = Alignment(wrap_text=True)

                        # Apply borders
                        for col in range(1, 11):
                            ws.cell(row=row_num, column=col).border = thin_border

                        row_num += 1
                else:
                    ws.cell(row=row_num, column=1, value=v['vendor'])
                    ws.cell(row=row_num, column=2, value=v['product'])
                    ws.cell(row=row_num, column=3, value=v['version'])
                    ws.cell(row=row_num, column=4, value=device_count)
                    ws.cell(row=row_num, column=5, value=device_list)
                    ws.cell(row=row_num, column=10, value='No CVEs found')
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border
                    row_num += 1

            # Column widths
            ws.column_dimensions['A'].width = 12  # Vendor
            ws.column_dimensions['B'].width = 12  # Product
            ws.column_dimensions['C'].width = 18  # Version
            ws.column_dimensions['D'].width = 12  # Device Count
            ws.column_dimensions['E'].width = 35  # Affected Devices
            ws.column_dimensions['F'].width = 18  # CVE ID
            ws.column_dimensions['G'].width = 12  # Severity
            ws.column_dimensions['H'].width = 10  # CVSS
            ws.column_dimensions['I'].width = 12  # Published
            ws.column_dimensions['J'].width = 80  # Description

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Auto-filter
            ws.auto_filter.ref = f"A1:J{row_num - 1}"

            wb.save(filepath)

            total_rows = row_num - 2
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {total_rows} rows for {len(versions)} versions to:\n{filepath}"
            )

        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
            logger.exception("Export error")

    def _export_devices_report(self):
        """Export device-centric vulnerability report to Excel"""
        if not hasattr(self, '_loaded_platforms') or not self._loaded_platforms:
            QMessageBox.warning(self, "No Data",
                                "Load a CSV first to export device-focused report.")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Export Devices Vulnerability Report",
            f"devices_vulnerabilities_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Device Vulnerabilities"

            # Severity colors for count cells
            severity_fills = {
                'critical': PatternFill('solid', fgColor='DC2626'),
                'high': PatternFill('solid', fgColor='EA580C'),
                'medium': PatternFill('solid', fgColor='CA8A04'),
                'low': PatternFill('solid', fgColor='16A34A'),
            }
            white_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill('solid', fgColor='1F2937')
            header_font = Font(bold=True, color='FFFFFF')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Headers
            headers = ['Device', 'Platform', 'Vendor', 'Product', 'Version',
                       'Total CVEs', 'Critical', 'High', 'Medium', 'Low']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row_num = 2
            for p in self._loaded_platforms:
                # Get CVE counts for this platform
                if p.cpe_vendor and p.cpe_product and p.cpe_version:
                    cves = self.cache.get_cves_for_version(
                        p.cpe_vendor, p.cpe_product, p.cpe_version
                    )

                    # Count by severity
                    counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
                    for cve in cves:
                        sev = cve.get('severity', '').upper()
                        if sev in counts:
                            counts[sev] += 1

                    total = len(cves)
                else:
                    total = 0
                    counts = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}

                # Write a row for each device
                for device in p.device_names:
                    ws.cell(row=row_num, column=1, value=device)
                    ws.cell(row=row_num, column=2, value=p.raw)
                    ws.cell(row=row_num, column=3, value=p.cpe_vendor or p.vendor)
                    ws.cell(row=row_num, column=4, value=p.cpe_product or p.product)
                    ws.cell(row=row_num, column=5, value=p.cpe_version or p.version)
                    ws.cell(row=row_num, column=6, value=total)

                    # Severity counts with color coding if > 0
                    crit_cell = ws.cell(row=row_num, column=7, value=counts['CRITICAL'])
                    if counts['CRITICAL'] > 0:
                        crit_cell.fill = severity_fills['critical']
                        crit_cell.font = white_font
                    crit_cell.alignment = Alignment(horizontal='center')

                    high_cell = ws.cell(row=row_num, column=8, value=counts['HIGH'])
                    if counts['HIGH'] > 0:
                        high_cell.fill = severity_fills['high']
                        high_cell.font = white_font
                    high_cell.alignment = Alignment(horizontal='center')

                    med_cell = ws.cell(row=row_num, column=9, value=counts['MEDIUM'])
                    if counts['MEDIUM'] > 0:
                        med_cell.fill = severity_fills['medium']
                        med_cell.font = white_font
                    med_cell.alignment = Alignment(horizontal='center')

                    low_cell = ws.cell(row=row_num, column=10, value=counts['LOW'])
                    if counts['LOW'] > 0:
                        low_cell.fill = severity_fills['low']
                        low_cell.font = white_font
                    low_cell.alignment = Alignment(horizontal='center')

                    # Apply borders
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border

                    row_num += 1

                # If no device names, still write a row with count
                if not p.device_names and p.device_count:
                    ws.cell(row=row_num, column=1, value=f"({p.device_count} devices)")
                    ws.cell(row=row_num, column=2, value=p.raw)
                    ws.cell(row=row_num, column=3, value=p.cpe_vendor or p.vendor)
                    ws.cell(row=row_num, column=4, value=p.cpe_product or p.product)
                    ws.cell(row=row_num, column=5, value=p.cpe_version or p.version)
                    ws.cell(row=row_num, column=6, value=total)
                    ws.cell(row=row_num, column=7, value=counts['CRITICAL'])
                    ws.cell(row=row_num, column=8, value=counts['HIGH'])
                    ws.cell(row=row_num, column=9, value=counts['MEDIUM'])
                    ws.cell(row=row_num, column=10, value=counts['LOW'])
                    for col in range(1, 11):
                        ws.cell(row=row_num, column=col).border = thin_border
                    row_num += 1

            # Column widths
            ws.column_dimensions['A'].width = 25  # Device
            ws.column_dimensions['B'].width = 35  # Platform
            ws.column_dimensions['C'].width = 12  # Vendor
            ws.column_dimensions['D'].width = 12  # Product
            ws.column_dimensions['E'].width = 18  # Version
            ws.column_dimensions['F'].width = 12  # Total CVEs
            ws.column_dimensions['G'].width = 10  # Critical
            ws.column_dimensions['H'].width = 10  # High
            ws.column_dimensions['I'].width = 10  # Medium
            ws.column_dimensions['J'].width = 10  # Low

            # Freeze header row
            ws.freeze_panes = 'A2'

            # Auto-filter
            ws.auto_filter.ref = f"A1:J{row_num - 1}"

            wb.save(filepath)

            total_rows = row_num - 2
            QMessageBox.information(
                self, "Export Complete",
                f"Exported {total_rows} device vulnerability records to:\n{filepath}"
            )

        except ImportError:
            QMessageBox.critical(self, "Missing Dependency",
                                 "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
        except Exception as e:
            QMessageBox.critical(self, "Export Error", str(e))
            logger.exception("Export error")

    def _select_all_valid(self):
        """Select all rows with valid CPE mappings"""
        self.platform_table.clearSelection()
        selection = self.platform_table.selectionModel()

        for i, p in enumerate(self.platform_model.platforms):
            if p.cpe_vendor and p.cpe_product and p.cpe_version:
                idx = self.platform_model.index(i, 0)
                selection.select(idx, selection.SelectionFlag.Select | selection.SelectionFlag.Rows)

    def _on_platform_selected(self):
        """Handle platform selection change"""
        rows = set(idx.row() for idx in self.platform_table.selectedIndexes())
        if len(rows) == 1:
            row = list(rows)[0]
            p = self.platform_model.platforms[row]

            # Load CVEs for this version if synced
            if p.cpe_vendor and p.cpe_product and p.cpe_version:
                cves = self.cache.get_cves_for_version(p.cpe_vendor, p.cpe_product, p.cpe_version)
                self.cve_model.load_cves(cves)
                self.cve_version_label.setText(
                    f"CVEs for {p.cpe_vendor}:{p.cpe_product}:{p.cpe_version} ({len(cves)} found)"
                )

                if cves:
                    self.results_tabs.setCurrentIndex(2)  # Switch to CVE tab
            else:
                self.cve_model.load_cves([])
                self.cve_version_label.setText("Invalid CPE mapping - edit vendor/product/version fields")

    def _start_sync(self):
        """Start NVD sync for selected platforms"""
        rows = set(idx.row() for idx in self.platform_table.selectedIndexes())
        platforms = self.platform_model.get_selected_for_sync(list(rows))

        if not platforms:
            QMessageBox.warning(self, "No Valid Selections",
                                "Select platforms with valid CPE mappings to sync.")
            return

        force_sync = self.force_sync_cb.isChecked()

        # Filter out already-synced unless force is checked
        if not force_sync:
            to_sync = []
            skipped = 0
            for p in platforms:
                if not self.cache.is_version_synced(p.cpe_vendor, p.cpe_product, p.cpe_version):
                    to_sync.append(p)
                else:
                    skipped += 1
                    self.platform_model.update_status(p.raw, "synced")

            if skipped > 0:
                self.status_label.setText(f"Skipped {skipped} already-synced versions")

            platforms = to_sync

        if not platforms:
            QMessageBox.information(self, "Nothing to Sync",
                                    "All selected versions are already synced. Check 'Force Re-sync' to update them.")
            return

        # Save API key
        api_key = self.api_key_input.text().strip() or None
        if api_key:
            self.settings.setValue("nvd_api_key", api_key)

        # Update UI
        self.sync_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setVisible(True)
        self.progress.setMaximum(len(platforms))
        self.progress.setValue(0)

        # Mark all as syncing
        for p in platforms:
            self.platform_model.update_status(p.raw, "syncing")

        # Start worker
        self.worker = SyncWorker(self.db_path, platforms, api_key)
        self.worker.progress.connect(self._on_sync_progress)
        self.worker.version_complete.connect(self._on_version_complete)
        self.worker.finished_all.connect(self._on_sync_finished)
        self.worker.start()

    def _stop_sync(self):
        """Stop the sync worker"""
        if self.worker:
            self.worker.stop()

    def _on_sync_progress(self, current: int, total: int, message: str):
        """Handle sync progress update"""
        self.progress.setValue(current)
        self.status_label.setText(message)

    def _on_version_complete(self, raw: str, result: dict):
        """Handle single version sync complete"""
        status = result.get("status", "error")
        self.platform_model.update_status(raw, status)

    def _on_sync_finished(self, summary: dict):
        """Handle sync complete"""
        self.sync_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress.setVisible(False)

        self.status_label.setText(
            f"Sync complete: {summary['synced']} versions, "
            f"{summary['total_cves']} CVEs found, {summary['errors']} errors"
        )

        # Reconnect to see data written by worker thread
        self.cache.reconnect()

        self._update_summary()
        self._load_from_cache()  # Refresh cached versions table
        self.worker = None

    def _update_summary(self):
        """Update the summary text"""
        overall = self.cache.get_overall_summary()
        versions = self.cache.get_version_summary()

        text = "=== CVE Cache Summary ===\n\n"
        text += f"Total CVEs: {overall.get('total', 0)}\n"
        text += f"  Critical: {overall.get('critical', 0)}\n"
        text += f"  High: {overall.get('high', 0)}\n"
        text += f"  Medium: {overall.get('medium', 0)}\n"
        text += f"  Low: {overall.get('low', 0)}\n\n"

        if versions:
            text += "=== Synced Versions ===\n\n"
            for v in versions[:20]:
                text += f"{v['vendor']}:{v['product']}:{v['version']}\n"
                text += f"  CVEs: {v['cve_count']} (C:{v['critical_count']} H:{v['high_count']} M:{v['medium_count']} L:{v['low_count']})\n"
                text += f"  Synced: {v['last_synced'][:10] if v['last_synced'] else 'Never'}\n\n"

        self.summary_text.setText(text)

    def _add_custom_pattern(self):
        """Show dialog to add a custom platform pattern"""
        dialog = AddPatternDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            pattern = dialog.get_pattern()
            self.parser.save_custom_pattern(**pattern)

            # Re-parse current platforms
            if self.platform_model.platforms:
                for p in self.platform_model.platforms:
                    if p.confidence == "low":
                        new_p = self.parser.parse(p.raw)
                        if new_p.confidence == "high":
                            p.cpe_vendor = new_p.cpe_vendor
                            p.cpe_product = new_p.cpe_product
                            p.cpe_version = new_p.cpe_version
                            p.confidence = "high"

                self.platform_model.layoutChanged.emit()

            QMessageBox.information(self, "Pattern Added",
                                    f"Pattern saved to {self.parser.custom_patterns_path}")

    def _show_help(self):
        """Show help dialog"""
        dialog = SecurityHelpDialog(self)
        if self.theme_manager:
            # Apply theme to help dialog
            theme = self.theme_manager.theme
            dialog.setStyleSheet(f"""
                QDialog {{
                    background-color: {theme.bg_primary};
                    color: {theme.text_primary};
                }}
                QTextEdit {{
                    background-color: {theme.bg_secondary};
                    color: {theme.text_primary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                }}
                QPushButton {{
                    background-color: {theme.accent};
                    color: {theme.text_on_accent};
                    border: none;
                    border-radius: 6px;
                    padding: 8px 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {theme.accent_hover};
                }}
                QLabel {{
                    color: {theme.text_primary};
                }}
            """)
        dialog.exec()

    def apply_theme(self, theme):
        """Apply theme colors to widget"""
        # Store for later use
        if hasattr(theme, 'bg_primary'):
            self._current_theme = theme

            # Main widget styling
            self.setStyleSheet(f"""
                QWidget {{
                    background-color: {theme.bg_primary};
                    color: {theme.text_primary};
                }}

                QGroupBox {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 8px;
                    margin-top: 12px;
                    padding-top: 8px;
                    font-weight: bold;
                }}

                QGroupBox::title {{
                    subcontrol-origin: margin;
                    left: 12px;
                    padding: 0 8px;
                    color: {theme.text_primary};
                }}

                QPushButton {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    padding: 8px 16px;
                    color: {theme.text_primary};
                }}

                QPushButton:hover {{
                    border-color: {theme.accent};
                    color: {theme.accent};
                }}

                QPushButton:pressed {{
                    background-color: {theme.bg_hover};
                }}

                QPushButton:disabled {{
                    background-color: {theme.bg_disabled};
                    color: {theme.text_disabled};
                    border-color: {theme.border_dim};
                }}

                QLineEdit {{
                    background-color: {theme.bg_input};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    padding: 8px;
                    color: {theme.text_primary};
                }}

                QLineEdit:focus {{
                    border-color: {theme.accent};
                }}

                QTableView {{
                    background-color: {theme.bg_secondary};
                    alternate-background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    gridline-color: {theme.border_dim};
                    color: {theme.text_primary};
                }}

                QTableView::item:selected {{
                    background-color: {theme.bg_selected};
                    color: {theme.accent};
                }}

                QHeaderView::section {{
                    background-color: {theme.bg_tertiary};
                    color: {theme.text_primary};
                    border: none;
                    border-bottom: 1px solid {theme.border_dim};
                    padding: 8px;
                    font-weight: bold;
                }}

                QTabWidget::pane {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                }}

                QTabBar::tab {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-bottom: none;
                    border-top-left-radius: 6px;
                    border-top-right-radius: 6px;
                    padding: 8px 16px;
                    color: {theme.text_secondary};
                }}

                QTabBar::tab:selected {{
                    background-color: {theme.bg_secondary};
                    color: {theme.accent};
                }}

                QTabBar::tab:hover:!selected {{
                    color: {theme.text_primary};
                }}

                QTextEdit {{
                    background-color: {theme.bg_secondary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    color: {theme.text_primary};
                }}

                QProgressBar {{
                    background-color: {theme.bg_tertiary};
                    border: 1px solid {theme.border_dim};
                    border-radius: 6px;
                    text-align: center;
                    color: {theme.text_primary};
                }}

                QProgressBar::chunk {{
                    background-color: {theme.accent};
                    border-radius: 5px;
                }}

                QCheckBox {{
                    color: {theme.text_primary};
                    spacing: 8px;
                }}

                QCheckBox::indicator {{
                    width: 18px;
                    height: 18px;
                    border: 1px solid {theme.border_secondary};
                    border-radius: 4px;
                    background-color: {theme.bg_input};
                }}

                QCheckBox::indicator:checked {{
                    background-color: {theme.accent};
                    border-color: {theme.accent};
                }}

                QLabel {{
                    color: {theme.text_primary};
                }}

                QScrollBar:vertical {{
                    background-color: {theme.bg_primary};
                    width: 10px;
                    margin: 0;
                }}

                QScrollBar::handle:vertical {{
                    background-color: {theme.scrollbar_handle};
                    border-radius: 5px;
                    min-height: 20px;
                }}

                QScrollBar::handle:vertical:hover {{
                    background-color: {theme.scrollbar_hover};
                }}

                QScrollBar::add-line:vertical,
                QScrollBar::sub-line:vertical {{
                    height: 0px;
                }}
            """)

    def closeEvent(self, event):
        """Clean up on close"""
        if self.worker and self.worker.isRunning():
            self.worker.stop()
            self.worker.wait()
        self.cache.close()
        super().closeEvent(event)


class AddPatternDialog(QDialog):
    """Dialog for adding custom platform patterns"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Platform Pattern")
        self.setMinimumWidth(500)

        layout = QFormLayout(self)

        self.regex_input = QLineEdit()
        self.regex_input.setPlaceholderText(r"e.g., MyVendor\s+OS\s+(\d+\.\d+\.\d+)")
        layout.addRow("Regex Pattern:", self.regex_input)

        self.vendor_input = QLineEdit()
        self.vendor_input.setPlaceholderText("e.g., MyVendor")
        layout.addRow("Display Vendor:", self.vendor_input)

        self.product_input = QLineEdit()
        self.product_input.setPlaceholderText("e.g., MyOS")
        layout.addRow("Display Product:", self.product_input)

        self.cpe_vendor_input = QLineEdit()
        self.cpe_vendor_input.setPlaceholderText("e.g., myvendor (lowercase)")
        layout.addRow("CPE Vendor:", self.cpe_vendor_input)

        self.cpe_product_input = QLineEdit()
        self.cpe_product_input.setPlaceholderText("e.g., myos (lowercase)")
        layout.addRow("CPE Product:", self.cpe_product_input)

        note = QLabel("Note: The regex should have a capture group () around the version number.")
        note.setStyleSheet("color: gray; font-style: italic;")
        layout.addRow(note)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)

    def get_pattern(self) -> dict:
        return {
            "regex": self.regex_input.text(),
            "vendor": self.vendor_input.text(),
            "product": self.product_input.text(),
            "cpe_vendor": self.cpe_vendor_input.text().lower(),
            "cpe_product": self.cpe_product_input.text().lower(),
        }


# ============================================================================
# Standalone Entry Point
# ============================================================================

def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    # Dark mode option
    if "--dark" in sys.argv:
        from PyQt6.QtGui import QPalette, QColor
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.WindowText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Base, QColor(25, 25, 25))
        palette.setColor(QPalette.ColorRole.AlternateBase, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.Text, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Button, QColor(53, 53, 53))
        palette.setColor(QPalette.ColorRole.ButtonText, Qt.GlobalColor.white)
        palette.setColor(QPalette.ColorRole.Highlight, QColor(42, 130, 218))
        app.setPalette(palette)

    widget = SecurityWidget()
    widget.setWindowTitle("Secure Cartography - Security Analysis")
    widget.resize(1000, 700)
    widget.show()

    sys.exit(app.exec())


if __name__ == "__main__":
    main()